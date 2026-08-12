"""
Import / re-sync Seema Enterprise inventory Excel into Category + Product (+ Supplier).

Excel PURCHASE/SELLING RATE WITHOUT GST are GST-exclusive.
Stored as purchase_price / selling_price (Without GST).
With-GST fields (purchase_price_with_gst / selling_price_with_gst) are set to 0.

Usage:
  python manage.py import_seema_inventory "C:\\path\\to\\file.xlsx" [--owner-id 4] [--dry-run] [--update-prices]
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from inventory.models import (
    Brand,
    Category,
    Product,
    StockMovement,
    Supplier,
    apply_stock_movement,
)

CATEGORY_RULES = [
    (r'\bDB\b|DISTRIBUTION|DB BOARD', 'Distribution Boards', '#6366f1'),
    (
        r'\bMCB\b|\bRCCB\b|\bELCB\b|\bTPN\b|\bDP\b\s*RENOVE|\bSP\b\s*(?:RENOVE|ARMOR|MCB)|'
        r'\d+\s*AMP\s*(?:SP|DP|TPN)|POLE MCB|MCB BOX',
        'MCBs & Protection',
        '#ef4444',
    ),
    (r'\bLED\b|\bLIGHT\b|\bBULB\b|\bT5\b|STREET LIGHT|NIGHT LAMP', 'LED & Lighting', '#f59e0b'),
    (r'\bCABLE\b|\bWIRE\b|INDOLEX|SQ\s*MM', 'Wires & Cables', '#0ea5e9'),
    (
        r'SURFACE BOX|SURFACE BOARD|\bBOX\b|BACK BOX|METAL BOX|FANBOX',
        'Boxes & Enclosures',
        '#8b5cf6',
    ),
    (r'\bPLATE\b|AIR PLATE', 'Modular Plates', '#14b8a6'),
    (
        r'REGULATOR|SWITCH|1 WAY|2 WAY|DOLLY|BELL\s*PUSH|BELLPUSH|MODULAR \d',
        'Switches & Regulators',
        '#22c55e',
    ),
    (
        r'SOCKET|PIN TOP|MULTIPLUG|3 PIN|WITHOUT SHUTTER|EXTENSION',
        'Sockets & Plugs',
        '#3b82f6',
    ),
    (r'\bBELL\b|DING DONG|ALARM|SENSOR', 'Bells & Alarms', '#ec4899'),
    (r'\bFAN\b|BLDC|PEDESTIAL|TABLE FAN', 'Fans & Accessories', '#a855f7'),
    (
        r'\bPIPE\b|\bPVC\b|\bGI\b|\bP3\b|\bMM\b.*(?:PIPE|P3)|RAPID\s*(?:TEE|COUPLER|JUNCTION)|'
        r'\bTEE\b|\bCOUPLER\b|JUNCTION|CLAMP',
        'Conduit & Pipes',
        '#64748b',
    ),
    (r'\bHEATER\b|\bJAGUAR\b|RADIUS HEATER', 'Heaters & Appliances', '#f97316'),
    (
        r'CONNECTOR|CABLE TIE|\bTIE\b|HOLDER|ANGLE|MFD|CAPACITOR|FASTNER|INDICATOR|'
        r'TESTER|MALE FEMALE|COMBINED|SKYLAB',
        'Accessories & Hardware',
        '#78716c',
    ),
]

DEFAULT_CATEGORY = ('General Electrical', '#0ea5e9')
CATEGORY_COLORS = {name: color for _, name, color in CATEGORY_RULES}
CATEGORY_COLORS[DEFAULT_CATEGORY[0]] = DEFAULT_CATEGORY[1]
GST_RATE = Decimal('18')

# Company / brand names often appear in the product title
BRAND_RULES = [
    (r'\bFYBROS\b', 'Fybros'),
    (r'\bRENOVE\b', 'Renove'),
    (r'\bARMOR\b', 'Armor'),
    (r'\bHYBRID\b', 'Hybrid'),
    (r'\bAURA\b', 'Aura'),
    (r'\bIKON\b', 'Ikon'),
    (r'\bGLOW\b', 'Glow'),
    (r'\bBELLA\b', 'Bella'),
    (r'\bAERIS\b', 'Fybros'),
    (r'\bRAZE\b', 'Raze'),
    (r'\bEME\b', 'Eme'),
    (r'\bPARROT\b', 'Parrot'),
    (r'\bAIR-?1\b|CEILING FAN', 'Air-1'),
    # Seema Enterprise Excel brands
    (r'\bINDOLEX\b', 'Indolex'),
    (r'\bFUTURE\b', 'Future'),
    (r'\bORASLE\b', 'Orasle'),
    (r'\bRAPID\b', 'Rapid'),
    (r'\bLEXUR\b', 'Lexur'),
    (r'\bPAYAL\b', 'Payal'),
    (r'\bFANCON\b', 'Fancon'),
    (r'\bJAGUAR\b', 'Jaguar'),
    (r'\bFEVILEX\b', 'Fevilex'),
    (r'\bPANCHAM\b', 'Pancham'),
    (r'\bORIENTAL\b', 'Oriental'),
    (r'\bRADIUS\b', 'Radius'),
    (r'\bVINTEX\b', 'Vintex'),
    (r'\bSKYLAB\b', 'Skylab'),
    (r'\bEPNIX\b', 'Epnix'),
    (r'\bJ\s+ULTIMATE\b|\bULTIMATE PEDESTIAL\b', 'J Ultimate'),
]


def _dec(value, default='0'):
    if value is None or value == '':
        return Decimal(default)
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def infer_category(name: str) -> str:
    upper = name.upper()
    for pattern, cat_name, _ in CATEGORY_RULES:
        if re.search(pattern, upper):
            return cat_name
    return DEFAULT_CATEGORY[0]


def infer_brand(name: str) -> str:
    """Detect company/brand from product name."""
    upper = name.upper()
    for pattern, brand in BRAND_RULES:
        if re.search(pattern, upper):
            return brand
    return ''


def clean_supplier(raw) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text.isdigit():
        return None
    return text[:200]


def parse_purchase_date(raw):
    if raw is None or raw == '':
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = 'Import Seema Enterprise inventory Excel (rates -> Without GST; With GST = 0)'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str)
        parser.add_argument('--owner-id', type=int, default=None)
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument(
            '--update-prices',
            action='store_true',
            help='Update existing products prices/stock from Excel instead of skipping',
        )

    def handle(self, *args, **options):
        path = Path(options['excel_path'])
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required: pip install openpyxl') from exc

        User = get_user_model()
        owner_id = options['owner_id']
        if owner_id:
            owner = User.objects.filter(pk=owner_id).first()
        else:
            owner = User.objects.order_by('id').first()
        if not owner:
            raise CommandError('No user found to own inventory')

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = []
        for excel_row in ws.iter_rows(min_row=4, max_col=8, values_only=True):
            purchase_date_raw, _sr, name, purchase_excl, sell_excl, stock, purchased_qty, supplier = excel_row
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            purchase_excl_d = _dec(purchase_excl)
            sell_excl_d = _dec(sell_excl)
            stock_d = _dec(stock)
            purchased_qty_d = _dec(purchased_qty)
            if (
                purchase_excl_d == 0
                and sell_excl_d == 0
                and stock_d == 0
                and purchase_excl is None
                and sell_excl is None
                and stock is None
            ):
                continue
            rows.append({
                'name': name,
                'purchase_date': parse_purchase_date(purchase_date_raw),
                'purchase_without_gst': purchase_excl_d,
                'sell_without_gst': sell_excl_d,
                'stock': stock_d,
                'purchased_qty': purchased_qty_d,
                'supplier': clean_supplier(supplier),
                'category': infer_category(name),
                'brand': infer_brand(name),
            })

        by_name = {}
        for row in rows:
            by_name[row['name'].casefold()] = row
        rows = list(by_name.values())

        cat_counts = {}
        for row in rows:
            cat_counts[row['category']] = cat_counts.get(row['category'], 0) + 1

        self.stdout.write(f'Owner id={owner.pk}')
        self.stdout.write(f'Products: {len(rows)} (Excel rates -> Without GST; With GST = 0)')
        for cat, count in sorted(cat_counts.items(), key=lambda x: (-x[1], x[0])):
            self.stdout.write(f'  - {cat}: {count}')

        if options['dry_run']:
            sample = rows[0]
            self.stdout.write(
                f'Sample {sample["name"]}: purchaseWithoutGst={sample["purchase_without_gst"]}, '
                f'sellWithoutGst={sample["sell_without_gst"]}, withGst=0'
            )
            self.stdout.write(self.style.WARNING('Dry run - no DB writes'))
            return

        created_cats = created_sups = created_prods = updated = skipped = stocked = 0

        with transaction.atomic():
            category_map = {}
            for cat_name in sorted(cat_counts):
                cat, was_created = Category.objects.get_or_create(
                    owner=owner,
                    name=cat_name,
                    defaults={
                        'description': 'Imported from Seema Enterprise inventory',
                        'color': CATEGORY_COLORS.get(cat_name, '#0ea5e9'),
                    },
                )
                category_map[cat_name] = cat
                if was_created:
                    created_cats += 1

            supplier_map = {}
            for row in rows:
                sname = row['supplier']
                if not sname or sname in supplier_map:
                    continue
                sup, was_created = Supplier.objects.get_or_create(
                    owner=owner,
                    name=sname,
                    defaults={},
                )
                supplier_map[sname] = sup
                if was_created:
                    created_sups += 1

            brand_map = {}
            for row in rows:
                bname = (row.get('brand') or '').strip()
                if not bname or bname.casefold() in brand_map:
                    continue
                brand, _ = Brand.objects.get_or_create(
                    owner=owner,
                    name=bname,
                    defaults={'color': '#6366f1'},
                )
                brand_map[bname.casefold()] = brand

            for row in rows:
                existing = Product.objects.filter(owner=owner, name__iexact=row['name']).first()
                brand_obj = brand_map.get((row.get('brand') or '').strip().casefold())
                if existing:
                    if options['update_prices']:
                        existing.purchase_price = row['purchase_without_gst']
                        existing.selling_price = row['sell_without_gst']
                        existing.purchase_price_with_gst = Decimal('0')
                        existing.selling_price_with_gst = Decimal('0')
                        existing.tax_rate = GST_RATE
                        existing.purchase_date = row['purchase_date'] or existing.purchase_date
                        existing.purchased_quantity = row['purchased_qty']
                        if brand_obj:
                            existing.brand = brand_obj
                        existing.category = category_map[row['category']]
                        if row['supplier']:
                            existing.supplier = supplier_map.get(row['supplier'])
                        existing.save()
                        updated += 1
                        if row['stock'] > 0 and Decimal(existing.stock_qty or 0) != row['stock']:
                            apply_stock_movement(
                                owner=owner,
                                product=existing,
                                movement_type=StockMovement.Type.ADJUST,
                                new_qty=row['stock'],
                                reason='Excel stock sync',
                                reference='SEEMA-XLSX',
                            )
                    else:
                        skipped += 1
                    continue

                product = Product.objects.create(
                    owner=owner,
                    name=row['name'][:200],
                    brand=brand_obj,
                    category=category_map[row['category']],
                    supplier=supplier_map.get(row['supplier']),
                    purchase_date=row['purchase_date'],
                    purchase_price=row['purchase_without_gst'],
                    selling_price=row['sell_without_gst'],
                    purchase_price_with_gst=Decimal('0'),
                    selling_price_with_gst=Decimal('0'),
                    tax_rate=GST_RATE,
                    stock_qty=Decimal('0'),
                    purchased_quantity=row.get('purchased_qty') or Decimal('0'),
                    status=Product.Status.ACTIVE,
                    description='Imported from Seema Enterprise inventory list',
                )
                created_prods += 1
                if row['stock'] > 0:
                    apply_stock_movement(
                        owner=owner,
                        product=product,
                        movement_type=StockMovement.Type.IN,
                        quantity=row['stock'],
                        reason='Opening stock (Excel import)',
                        reference='SEEMA-XLSX',
                    )
                    stocked += 1

        self.stdout.write(self.style.SUCCESS(
            f'Done. categories+={created_cats}, suppliers+={created_sups}, '
            f'products+={created_prods}, updated={updated}, stocked={stocked}, skipped={skipped}'
        ))
