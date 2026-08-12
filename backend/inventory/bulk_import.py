"""Parse CSV / Excel product uploads into create/update rows."""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction

from .models import (
    Brand,
    Category,
    Product,
    StockMovement,
    Supplier,
    apply_stock_movement,
)

# Reuse name→category/brand heuristics from the Seema CLI importer
from .management.commands.import_seema_inventory import (
    CATEGORY_COLORS,
    DEFAULT_CATEGORY,
    GST_RATE,
    infer_brand,
    infer_category,
)

STATUS_MAP = {
    'active': Product.Status.ACTIVE,
    'inactive': Product.Status.INACTIVE,
    'discontinued': Product.Status.DISCONTINUED,
}


def _dec(value, default='0'):
    if value is None or value == '':
        return Decimal(default)
    text = str(value).strip().replace(',', '')
    if not text:
        return Decimal(default)
    try:
        return Decimal(text)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _norm_key(key) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(key or '').strip().lower())


HEADER_ALIASES = {
    'name': {'name', 'productname', 'product'},
    'brand': {'brand', 'company', 'manufacturer'},
    'category': {'category'},
    'stock': {'stock', 'stockqty', 'qty'},
    'purchased_qty': {'purchasedquantity', 'purchasedqty'},
    'purchase_date': {'purchasedate'},
    # Prefer *with*gst / *without*gst before bare purchaserate / sellingrate
    'purchase_with': {
        'purchasewithgst',
        'purchasepricewithgst',
        'purchaseratewithgst',
    },
    'purchase_without': {
        'purchasewithoutgst',
        'purchasepricewithoutgst',
        'purchaseratewithoutgst',
        'purchaseprice',
        'purchaserate',
    },
    'sell_with': {
        'sellingwithgst',
        'sellingpricewithgst',
        'sellingratewithgst',
    },
    'sell_without': {
        'sellingwithoutgst',
        'sellingpricewithoutgst',
        'sellingratewithoutgst',
        'sellingprice',
        'sellingrate',
    },
    'gst_rate': {'gstrate', 'taxrate', 'gst'},
    'status': {'status'},
    'supplier': {'supplier', 'purchasedfrom'},
}

# Longer / more specific aliases win when multiple could match
_ALIAS_LOOKUP = sorted(
    ((alias, field) for field, aliases in HEADER_ALIASES.items() for alias in aliases),
    key=lambda x: (-len(x[0]), x[0]),
)


def _map_headers(raw_headers):
    """Map row keys → canonical field names (specific GST headers first)."""
    mapping = {}
    for raw in raw_headers:
        if raw is None or str(raw).strip() == '':
            continue
        nk = _norm_key(raw)
        for alias, field in _ALIAS_LOOKUP:
            if nk == alias and field not in mapping:
                mapping[field] = raw
                break
    return mapping


def parse_purchase_date(raw):
    if raw is None or raw == '':
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _row_from_mapped(mapped: dict, get) -> dict | None:
    name = str(get('name') or '').strip()
    if not name:
        return None
    brand = str(get('brand') or '').strip()
    category = str(get('category') or '').strip()
    supplier = str(get('supplier') or '').strip()
    if supplier.isdigit():
        supplier = ''

    present = set(mapped.keys())
    purchase_without = _dec(get('purchase_without')) if 'purchase_without' in present else None
    purchase_with = _dec(get('purchase_with')) if 'purchase_with' in present else None
    sell_without = _dec(get('sell_without')) if 'sell_without' in present else None
    sell_with = _dec(get('sell_with')) if 'sell_with' in present else None
    stock = _dec(get('stock')) if 'stock' in present else None
    purchased_qty = _dec(get('purchased_qty')) if 'purchased_qty' in present else None
    gst_raw = get('gst_rate') if 'gst_rate' in present else None
    tax_rate = _dec(gst_raw, str(GST_RATE)) if gst_raw not in (None, '') else (
        GST_RATE if 'gst_rate' in present else None
    )
    status_raw = str(get('status') or 'active').strip().lower() if 'status' in present else None
    status = STATUS_MAP.get(status_raw, Product.Status.ACTIVE) if status_raw else None

    if not brand:
        brand = infer_brand(name)
    if not category:
        category = infer_category(name)

    return {
        'name': name[:200],
        'brand': brand[:100] if brand else '',
        'category': category or DEFAULT_CATEGORY[0],
        'supplier': supplier[:200] if supplier else '',
        'purchase_date': parse_purchase_date(get('purchase_date')) if 'purchase_date' in present else None,
        'purchase_without': purchase_without,
        'purchase_with': purchase_with,
        'sell_without': sell_without,
        'sell_with': sell_with,
        'tax_rate': tax_rate,
        'stock': stock,
        'purchased_qty': purchased_qty,
        'status': status,
        'present': present,
    }


def _merge_rows(prev: dict, nxt: dict) -> dict:
    """Merge duplicate-name rows: keep non-empty prices; prefer latest stock/qty/meta."""
    merged = {**prev}
    present = set(prev.get('present') or set()) | set(nxt.get('present') or set())
    for key in (
        'purchase_without',
        'purchase_with',
        'sell_without',
        'sell_with',
        'tax_rate',
        'stock',
        'purchased_qty',
        'status',
        'purchase_date',
        'supplier',
        'brand',
        'category',
    ):
        new_val = nxt.get(key)
        old_val = prev.get(key)
        if new_val is None or new_val == '':
            continue
        # Prefer non-zero prices when previous was empty/zero
        if key in ('purchase_without', 'purchase_with', 'sell_without', 'sell_with'):
            if old_val in (None, Decimal('0')) and new_val != Decimal('0'):
                merged[key] = new_val
            elif new_val is not None:
                merged[key] = new_val
        else:
            merged[key] = new_val
    if nxt.get('brand'):
        merged['brand'] = nxt['brand']
    if nxt.get('category'):
        merged['category'] = nxt['category']
    merged['present'] = present
    return merged


def _dedupe_by_name(rows: list[dict]) -> tuple[list[dict], int]:
    by_name = {}
    collapsed = 0
    for row in rows:
        key = row['name'].casefold()
        if key in by_name:
            collapsed += 1
            by_name[key] = _merge_rows(by_name[key], row)
        else:
            by_name[key] = row
    return list(by_name.values()), collapsed


def parse_csv_bytes(data: bytes) -> tuple[list[dict], int, int]:
    text = data.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError('CSV has no header row')
    mapping = _map_headers(reader.fieldnames)
    if 'name' not in mapping:
        raise ValueError('CSV must include a Name / Product Name column')

    rows = []
    for raw in reader:
        def get(field):
            key = mapping.get(field)
            return raw.get(key) if key else None

        parsed = _row_from_mapped(mapping, get)
        if parsed:
            rows.append(parsed)
    source = len(rows)
    deduped, collapsed = _dedupe_by_name(rows)
    return deduped, source, collapsed


def parse_xlsx_bytes(data: bytes) -> tuple[list[dict], int, int]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError('openpyxl is required for Excel import') from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx = None
    headers = []
    for r in range(1, min(6, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        mapping = _map_headers(vals)
        if 'name' in mapping:
            header_row_idx = r
            headers = vals
            break

    if header_row_idx is None:
        raise ValueError('Excel must include a Name / Product Name header')

    mapping = _map_headers(headers)
    col_index = {field: headers.index(raw) + 1 for field, raw in mapping.items()}

    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        def get(field, _r=r):
            idx = col_index.get(field)
            return ws.cell(_r, idx).value if idx else None

        parsed = _row_from_mapped(mapping, get)
        if parsed:
            rows.append(parsed)
    source = len(rows)
    deduped, collapsed = _dedupe_by_name(rows)
    return deduped, source, collapsed


def parse_upload(filename: str, data: bytes) -> tuple[list[dict], dict]:
    lower = (filename or '').lower()
    if lower.endswith('.csv'):
        rows, source, collapsed = parse_csv_bytes(data)
    elif lower.endswith('.xlsx') or lower.endswith('.xlsm'):
        rows, source, collapsed = parse_xlsx_bytes(data)
    elif data[:2] == b'PK':
        rows, source, collapsed = parse_xlsx_bytes(data)
    else:
        rows, source, collapsed = parse_csv_bytes(data)
    return rows, {
        'sourceRows': source,
        'uniqueProducts': len(rows),
        'duplicatesCollapsed': collapsed,
    }


def import_product_rows(owner, rows: list[dict], *, update_existing: bool = True) -> dict:
    """Create/update products for owner. Returns summary counts + errors."""
    created = updated = skipped = stocked = 0
    errors: list[str] = []

    with transaction.atomic():
        category_map = {}
        brand_map = {}
        supplier_map = {}

        for row in rows:
            cat_name = row['category']
            if cat_name not in category_map:
                cat, _ = Category.objects.get_or_create(
                    owner=owner,
                    name=cat_name,
                    defaults={
                        'description': 'Created from product import',
                        'color': CATEGORY_COLORS.get(cat_name, '#0ea5e9'),
                    },
                )
                category_map[cat_name] = cat

            bname = (row.get('brand') or '').strip()
            if bname and bname.casefold() not in brand_map:
                brand, _ = Brand.objects.get_or_create(
                    owner=owner,
                    name=bname,
                    defaults={'color': '#6366f1'},
                )
                brand_map[bname.casefold()] = brand

            sname = (row.get('supplier') or '').strip()
            if sname and sname not in supplier_map:
                sup, _ = Supplier.objects.get_or_create(
                    owner=owner,
                    name=sname,
                    defaults={},
                )
                supplier_map[sname] = sup

        for row in rows:
            try:
                brand_obj = brand_map.get((row.get('brand') or '').strip().casefold())
                present = row.get('present') or set()
                existing = Product.objects.filter(owner=owner, name__iexact=row['name']).first()
                if existing:
                    if not update_existing:
                        skipped += 1
                        continue
                    # Only overwrite fields that exist in this upload (so With/Without GST files merge)
                    if 'purchase_without' in present and row['purchase_without'] is not None:
                        existing.purchase_price = row['purchase_without']
                    if 'sell_without' in present and row['sell_without'] is not None:
                        existing.selling_price = row['sell_without']
                    if 'purchase_with' in present and row['purchase_with'] is not None:
                        existing.purchase_price_with_gst = row['purchase_with']
                    if 'sell_with' in present and row['sell_with'] is not None:
                        existing.selling_price_with_gst = row['sell_with']
                    if 'tax_rate' in present and row['tax_rate'] is not None:
                        existing.tax_rate = row['tax_rate']
                    elif existing.tax_rate in (None, Decimal('0')):
                        existing.tax_rate = GST_RATE
                    if 'purchased_qty' in present and row['purchased_qty'] is not None:
                        existing.purchased_quantity = row['purchased_qty']
                    if 'status' in present and row['status'] is not None:
                        existing.status = row['status']
                    if row.get('purchase_date'):
                        existing.purchase_date = row['purchase_date']
                    if brand_obj:
                        existing.brand = brand_obj
                    existing.category = category_map[row['category']]
                    if row.get('supplier'):
                        existing.supplier = supplier_map.get(row['supplier'])
                    existing.save()
                    updated += 1
                    if 'stock' in present and row['stock'] is not None:
                        if Decimal(existing.stock_qty or 0) != row['stock']:
                            apply_stock_movement(
                                owner=owner,
                                product=existing,
                                movement_type=StockMovement.Type.ADJUST,
                                new_qty=row['stock'],
                                reason='Bulk import stock sync',
                                reference='BULK-IMPORT',
                            )
                            stocked += 1
                    continue

                product = Product.objects.create(
                    owner=owner,
                    name=row['name'],
                    brand=brand_obj,
                    category=category_map[row['category']],
                    supplier=supplier_map.get(row.get('supplier') or ''),
                    purchase_date=row.get('purchase_date'),
                    purchase_price=row['purchase_without'] if row['purchase_without'] is not None else Decimal('0'),
                    selling_price=row['sell_without'] if row['sell_without'] is not None else Decimal('0'),
                    purchase_price_with_gst=row['purchase_with'] if row['purchase_with'] is not None else Decimal('0'),
                    selling_price_with_gst=row['sell_with'] if row['sell_with'] is not None else Decimal('0'),
                    tax_rate=row['tax_rate'] if row['tax_rate'] is not None else GST_RATE,
                    stock_qty=Decimal('0'),
                    purchased_quantity=row['purchased_qty'] if row['purchased_qty'] is not None else Decimal('0'),
                    status=row['status'] or Product.Status.ACTIVE,
                    description='Imported via bulk upload',
                )
                created += 1
                if row.get('stock') is not None and row['stock'] > 0:
                    apply_stock_movement(
                        owner=owner,
                        product=product,
                        movement_type=StockMovement.Type.IN,
                        quantity=row['stock'],
                        reason='Opening stock (bulk import)',
                        reference='BULK-IMPORT',
                    )
                    stocked += 1
            except Exception as exc:  # noqa: BLE001 — collect row errors
                errors.append(f'{row.get("name", "?")}: {exc}')

    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'stocked': stocked,
        'total': len(rows),
        'errors': errors[:20],
    }
